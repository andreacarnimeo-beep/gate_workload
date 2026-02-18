import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from datetime import datetime, date

st.set_page_config(page_title="Gate Workload • EasyMag", layout="wide")

# ----------------------------
# Helpers
# ----------------------------
MARKER_PATTERN = re.compile(r"Numero\s+di\s+Operazioni", re.IGNORECASE)

def _find_marker_row(xlsx_bytes: bytes) -> int | None:
    """Return 0-based row index of marker in column A, or None."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True)):
        v = row[0]
        if isinstance(v, str) and MARKER_PATTERN.search(v):
            return i
    return None

def read_easymag_pivot(xlsx_bytes: bytes) -> pd.DataFrame:
    """
    Reads an EasyMag 'Statistica' export shaped like:
    Row1: Giro/Data | <Giro1> | <Giro2> | ...
    Rows: date values
    Row: Tot:
    Then a marker line "(*) Numero di Operazioni" and the table repeats.
    Returns LONG df with columns: Data (datetime or NaT), Giro (str), Valore (float).
    """
    marker0 = _find_marker_row(xlsx_bytes)
    # marker0 is 0-based index, pandas uses nrows count
    nrows = marker0 if marker0 is not None else None

    df = pd.read_excel(BytesIO(xlsx_bytes), header=0, nrows=nrows)
    if df.empty:
        return pd.DataFrame(columns=["Data", "Giro", "Valore"])

    first_col = df.columns[0]
    # drop fully empty rows
    df = df.dropna(how="all")
    # Ensure giro columns are strings
    df.columns = [str(c).strip() for c in df.columns]

    # Melt pivot to long
    long = df.melt(id_vars=[first_col], var_name="Giro", value_name="Valore")
    long.rename(columns={first_col: "DataRaw"}, inplace=True)

    # Clean Giro
    long["Giro"] = long["Giro"].astype(str).str.strip()

    # Parse dates; keep Tot row as NaT (but will be excluded from daily trend)
    def parse_date(x):
        if isinstance(x, str) and x.strip().lower().startswith("tot"):
            return pd.NaT
        return pd.to_datetime(x, errors="coerce")
    long["Data"] = long["DataRaw"].apply(parse_date)
    long.drop(columns=["DataRaw"], inplace=True)

    # Coerce numeric
    long["Valore"] = pd.to_numeric(long["Valore"], errors="coerce").fillna(0)

    # Drop blank giro labels if any
    long = long[long["Giro"].ne("")].copy()

    return long[["Data", "Giro", "Valore"]]

def read_gate_map(gate_file: BytesIO) -> pd.DataFrame:
    g = pd.read_excel(gate_file)
    # In your file: Giro is column B, Gate is column J; but we rely on headers if present.
    # If headers differ, fallback to positional.
    cols = {c.lower(): c for c in g.columns}
    giro_col = cols.get("giro", g.columns[1] if len(g.columns) > 1 else g.columns[0])
    gate_col = cols.get("gate", g.columns[9] if len(g.columns) > 9 else g.columns[-1])

    out = g[[giro_col, gate_col]].copy()
    out.columns = ["Giro", "Gate"]
    out["Giro"] = out["Giro"].astype(str).str.strip()
    # Gate may be numeric floats like 4.0 -> convert to int-like string for display
    out["Gate"] = out["Gate"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    out.loc[out["Gate"].isin(["nan", "None", ""]), "Gate"] = "NON_ASSEGNATO"
    return out.dropna(subset=["Giro"])

def label_metric(file_name: str, total_value: float) -> str:
    n = (file_name or "").lower()
    if "righe" in n or "prelev" in n:
        return "Righe"
    if "colli" in n:
        return "Colli"
    # fallback will be decided later comparing totals
    return "Auto"

def build_fact(long_df: pd.DataFrame, gate_map: pd.DataFrame, metric_name: str) -> pd.DataFrame:
    fact = long_df.copy()
    fact["Giro"] = fact["Giro"].astype(str).str.strip()
    fact["Metrica"] = metric_name
    fact = fact.merge(gate_map, on="Giro", how="left")
    fact["Gate"] = fact["Gate"].fillna("NON_ASSEGNATO")
    return fact

def safe_date_bounds(df: pd.DataFrame):
    d = df["Data"].dropna()
    if d.empty:
        today = pd.Timestamp(date.today())
        return today, today
    return d.min().normalize(), d.max().normalize()

# ----------------------------
# UI
# ----------------------------
st.title("Gate Workload • EasyMag")

left = st.sidebar
left.header("1) Caricamento file")

gate_file = left.file_uploader("GATE.xlsx (mappa Giro → Gate)", type=["xlsx"], key="gate")
f1 = left.file_uploader("Export EasyMag #1 (righe o colli)", type=["xlsx"], key="f1")
f2 = left.file_uploader("Export EasyMag #2 (righe o colli)", type=["xlsx"], key="f2")

left.header("2) Impostazioni")
top_n = left.slider("Top N giri (vista dettaglio)", 5, 50, 15)
thr_gate = left.slider("Alert: Gate > X% del totale", 5, 90, 40)
thr_ratio = left.slider("Alert: Colli/100 righe > soglia", 10, 500, 120)

if not (gate_file and f1 and f2):
    st.info("Carica **GATE.xlsx** e i **due export EasyMag** per vedere la dashboard.")
    st.stop()

gate_map = read_gate_map(gate_file)

# Read + truncate + pivot->long
long1 = read_easymag_pivot(f1.getvalue())
long2 = read_easymag_pivot(f2.getvalue())

tot1 = float(long1.loc[long1["Data"].notna(), "Valore"].sum())
tot2 = float(long2.loc[long2["Data"].notna(), "Valore"].sum())

lab1 = label_metric(f1.name, tot1)
lab2 = label_metric(f2.name, tot2)

# If both Auto, guess: larger total = Righe
if lab1 == "Auto" and lab2 == "Auto":
    if tot1 >= tot2:
        lab1, lab2 = "Righe", "Colli"
    else:
        lab1, lab2 = "Colli", "Righe"
# If one Auto, assign the other
if lab1 == "Auto" and lab2 != "Auto":
    lab1 = "Colli" if lab2 == "Righe" else "Righe"
if lab2 == "Auto" and lab1 != "Auto":
    lab2 = "Colli" if lab1 == "Righe" else "Righe"

fact = pd.concat([
    build_fact(long1, gate_map, lab1),
    build_fact(long2, gate_map, lab2),
], ignore_index=True)

# Date filter
dmin, dmax = safe_date_bounds(fact)
colA, colB, colC = st.columns([2,2,2])
with colA:
    metric_view = st.radio("Metrica", ["Righe", "Colli"], horizontal=True)
with colB:
    start = st.date_input("Dal", value=dmin.date(), min_value=dmin.date(), max_value=dmax.date())
with colC:
    end = st.date_input("Al", value=dmax.date(), min_value=dmin.date(), max_value=dmax.date())

start_ts = pd.Timestamp(start)
end_ts = pd.Timestamp(end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

# Filters Gate/Giri
all_gates = sorted(fact["Gate"].dropna().unique().tolist())
all_giri = sorted(fact["Giro"].dropna().unique().tolist())

with st.sidebar:
    st.header("3) Filtri")
    sel_gates = st.multiselect("Solo Gate selezionati", all_gates, default=all_gates)
    sel_giri = st.multiselect("Solo Giri selezionati", all_giri, default=all_giri)

f = fact[
    (fact["Metrica"] == metric_view) &
    (fact["Giro"].isin(sel_giri)) &
    (fact["Gate"].isin(sel_gates)) &
    (fact["Data"].between(start_ts, end_ts) | fact["Data"].isna())  # Tot rows have NaT; we keep but won't use for trends
].copy()

# Aggregate by Gate
gate_tot = (
    f[f["Data"].notna()]
    .groupby("Gate", as_index=False)["Valore"].sum()
    .sort_values("Valore", ascending=False)
)

total = gate_tot["Valore"].sum()
gate_tot["Perc"] = np.where(total > 0, gate_tot["Valore"] / total * 100, 0)

# ----------------------------
# Main dashboard: pie + stacked bars
# ----------------------------
c1, c2 = st.columns([1, 2], gap="large")

with c1:
    st.subheader("Peso Gate sul totale")
    import plotly.express as px
    fig = px.pie(gate_tot, names="Gate", values="Valore", hole=0.35)
    st.plotly_chart(fig, use_container_width=True)

with c2:
    st.subheader("Gate (stacked) → dettaglio giri")
    # Build stacked: top giri per gate
    d = (
        f[f["Data"].notna()]
        .groupby(["Gate", "Giro"], as_index=False)["Valore"].sum()
    )
    # keep top giri overall to reduce clutter
    top_giri_overall = (
        d.groupby("Giro")["Valore"].sum().sort_values(ascending=False).head(25).index.tolist()
    )
    d_plot = d[d["Giro"].isin(top_giri_overall)].copy()
    fig2 = px.bar(d_plot, x="Gate", y="Valore", color="Giro", barmode="stack")
    st.plotly_chart(fig2, use_container_width=True)

# ----------------------------
# Compare Righe vs Colli (single view)
# ----------------------------
st.subheader("Confronto Righe vs Colli per Gate (stesso periodo/filtri)")

base = fact[
    (fact["Giro"].isin(sel_giri)) &
    (fact["Gate"].isin(sel_gates)) &
    (fact["Data"].between(start_ts, end_ts) | fact["Data"].isna())
].copy()

cmp = (
    base[base["Data"].notna()]
    .groupby(["Gate", "Metrica"], as_index=False)["Valore"].sum()
)
cmp_piv = cmp.pivot_table(index="Gate", columns="Metrica", values="Valore", aggfunc="sum").fillna(0).reset_index()
if "Righe" not in cmp_piv.columns: cmp_piv["Righe"] = 0.0
if "Colli" not in cmp_piv.columns: cmp_piv["Colli"] = 0.0
cmp_piv["Colli/100 righe"] = np.where(cmp_piv["Righe"] > 0, cmp_piv["Colli"] / cmp_piv["Righe"] * 100, np.nan)

fig3 = px.bar(
    cmp_piv.sort_values("Righe", ascending=False),
    x="Gate",
    y=["Righe", "Colli"],
    barmode="group"
)
st.plotly_chart(fig3, use_container_width=True)
st.dataframe(cmp_piv.sort_values("Righe", ascending=False), use_container_width=True)

# ----------------------------
# Top Giri per Gate
# ----------------------------
st.subheader("Top giri per Gate (dettaglio)")
sel_gate_detail = st.selectbox("Seleziona Gate", options=sorted(gate_tot["Gate"].unique().tolist()))
top = (
    f[(f["Data"].notna()) & (f["Gate"] == sel_gate_detail)]
    .groupby("Giro", as_index=False)["Valore"].sum()
    .sort_values("Valore", ascending=False)
    .head(top_n)
)
fig4 = px.bar(top, x="Giro", y="Valore")
st.plotly_chart(fig4, use_container_width=True)

csv = top.to_csv(index=False).encode("utf-8")
st.download_button("Scarica CSV (Top giri)", data=csv, file_name=f"top_giri_{sel_gate_detail}_{metric_view}.csv", mime="text/csv")

# ----------------------------
# Trend giornaliero per Gate
# ----------------------------
st.subheader("Trend giornaliero per Gate")
trend = (
    f[f["Data"].notna()]
    .assign(Giorno=lambda x: x["Data"].dt.normalize())
    .groupby(["Giorno", "Gate"], as_index=False)["Valore"].sum()
)

sel_gates_trend = st.multiselect("Gate da mostrare nel trend", options=sorted(trend["Gate"].unique()), default=sorted(trend["Gate"].unique())[:5])
trend2 = trend[trend["Gate"].isin(sel_gates_trend)].copy()
fig5 = px.line(trend2, x="Giorno", y="Valore", color="Gate", markers=True)
st.plotly_chart(fig5, use_container_width=True)

# ----------------------------
# Alerts
# ----------------------------
st.subheader("Alert")
alerts = []

# Gate share alert (based on selected metric_view)
bad_share = gate_tot[gate_tot["Perc"] > thr_gate]
if not bad_share.empty:
    alerts.append(("Gate dominanti", bad_share[["Gate", "Perc", "Valore"]]))

# Colli/100 righe alert
bad_ratio = cmp_piv[cmp_piv["Colli/100 righe"] > thr_ratio].copy()
if not bad_ratio.empty:
    alerts.append(("Rapporto colli/righe alto", bad_ratio[["Gate", "Colli/100 righe", "Righe", "Colli"]]))

if not alerts:
    st.success("Nessun alert per le soglie impostate.")
else:
    for title, df_a in alerts:
        st.warning(title)
        st.dataframe(df_a, use_container_width=True)
